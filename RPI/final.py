
import time
from picamera.array import PiRGBArray
from picamera import PiCamera

import openpyxl
import serial
import sys
import time
import serial.tools.list_ports
import RPi.GPIO as GPIO
import cv2
import imutils
import numpy as np
import os
import importlib
#used to fetch details from the exel
# input format : exelLocation and the coordinates where to append things
def fetchDetailsFromExel(exel_name,coordinates):
    wb = openpyxl.load_workbook(exel_name)
    j=2
    while j<=wb['Sheet1'].max_row:
        name=(wb['Sheet1'].cell(j,1).value)
        x=(int(wb['Sheet1'].cell(j,2).value))
        y=(int(wb['Sheet1'].cell(j,3).value))
        a=int(wb['Sheet1'].cell(j,4).value)
        val=(x,y,a,name)
        coordinates.append(val)
        j=j+1
    return
#convert the gerber file data according to the machine data
def cnvtData(coordinates):
    j=0
    while j<len(coordinates):
        name=coordinates[j][3]
        x=(int((coordinates[j][0]*10))+550)*10
        y=(int((coordinates[j][1]*10))+2550)*10
        a=coordinates[j][2]
        val=(x,y,a,name)
        coordinates[j]=val
        j=j+1
    return

def setup():
    GPIO.cleanup()
    GPIO.setwarnings(False)
    GPIO.setmode(GPIO.BOARD)
    GPIO.setup(vacumn_pin, GPIO.OUT)
    GPIO.setup(led_pin, GPIO.OUT)
    GPIO.output(led_pin, GPIO.HIGH)
    GPIO.output(vacumn_pin, GPIO.HIGH)
    global arduino,f
    try:
        f=open("opencv.txt","r")
    except IOError:
       print ('No file with opencv found in the folder')
       exit()
    files = f.readlines()
    print (os.getcwd())
    for File in files:
        try:
            importlib.import_module(File)
        except ImportError:
            print "no "+File+" found in this directory"
            exit()
    arduino = serial.Serial('/dev/ttyUSB0',9600)
    arduino.baudrate=9600
    arduino.write('r')#reset arduino
    reci=arduino.readline()
    if(reci.find(unicode("Available"))==0):
        print("Pick and place machine found")
        arduino.write ('i')
        reci=arduino.readline()
        if(reci.find('Init Ended')==0):
            print("initialed at 0,0,0")
            arduino.write('e')#enable motors
            reci=arduino.readline()
            if(reci.find('Ended')):
                print ('initialed')
    return

def checkOuput():
    global arduino
    cont=True
    while cont:
        reci=arduino.readline()
        if(reci.find('Ended')==0):
            cont=False
    return

def main():
    global no_of_com,pick_coordinates,drop_coordinates,source,camera1
    for i in range(no_of_com):
        print ('picking'+pick_coordinates[i][3])
        x,y=pick_coordinates[i][0],pick_coordinates[i][1]
        arduino.write('x '+str(x));
        checkOuput()
        arduino.write('y '+str(y))
        checkOuput()
        arduino.write('z '+str(0))
        checkOuput()
        GPIO.output(vacumn_pin,GPIO.LOW)
        arduino.write('z '+str(45))
        checkOuput()
        arduino.write('c')
        checkOuput()
        GPIO.output(led_pin,GPIO.LOW)
        (extraX,extraY)=offsetAngleAndCentre(drop_coordinates[i][3],drop_coordinates[i][2],source,camera1)
        arduino.write('z '+str(45))
        checkOuput()
        GPIO.output(led_pin,GPIO.HIGH)
        print ('placing '+drop_coordinates[i][3])
        x,y=drop_coordinates[i][0],drop_coordinates[i][1]
        arduino.write('x '+str(x))
        checkOuput()
        arduino.write('y '+str(y))
        checkOuput()
        arduino.write('z 5')
        GPIO.output(vacumn_pin,GPIO.HIGH)
        checkOuput()
        arduino.write('z 0')
        checkOuput()
        arduino.write('z 45')
        checkOuput()
    arduino.write('d')
    checkOuput()
    print('completed')
    return

def offsetAngleAndCentre(name,deg,s,c):
    f=open("opencv.txt","r")
    files = f.readlines()
    done=False
    for x in files:
        x=importlib.import_module(x)
        if x.checkAvailable(name):
            (extraX,extraY)=x.offsetAngle(deg,arduino,s,c)
            done=True
            break
    if (done is True):
        return (extraX,extraY)
    else:
        print("can't correct")
        exit()

if __name__ == "__main__":
    drop_coordinates=[]
    pick_coordinates=[]
    no_of_com = 0

    vacumn_pin = 12
    led_pin = 11

    camera1 = PiCamera()
    camera1.resolution = (1920, 1080)
    camera1.framerate = 20
    source = PiRGBArray(camera1, size=(1920, 1080))

    arduino = None

    f= None

    centroid_file=sys.argv[1]
    print('Fetching centroid_file from '+centroid_file)

    pickup_file = sys.argv[2]
    print('Fetching pickup file from '+pickup_file)

    user_input=raw_input("Would you like to continue <y/n>")
    if (user_input.lower()=='n'):
        exit()
    fetchDetailsFromExel(centroid_file,drop_coordinates);no_of_com = len(drop_coordinates);cnvtData(drop_coordinates)
    fetchDetailsFromExel(pickup_file,pick_coordinates)
    print('Recods fetched and total number of components to be palced is ',no_of_com)
    setup()
    try:
      main()
    except KeyboardInterrupt:
        cv2.destroyAllWindows()
        GPIO.output(led_pin, GPIO.HIGH)
        GPIO.output(vacumn_pin, GPIO.HIGH)
        GPIO.cleanup()
        arduino.write('d')
